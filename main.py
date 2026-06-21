from pathlib import Path
from io import StringIO
from datetime import datetime
import pandas as pd


def numero_br(valor):
    if pd.isna(valor):
        return 0.0

    valor = str(valor).strip()

    if valor == "":
        return 0.0

    valor = valor.replace(".", "")
    valor = valor.replace(",", ".")

    try:
        return float(valor)
    except:
        return 0.0


def ler_txt_a_partir_do_cabecalho(caminho: Path) -> pd.DataFrame:
    with open(caminho, encoding="cp1252") as arquivo:
        linhas = arquivo.readlines()

    linha_cabecalho = None

    for i, linha in enumerate(linhas):
        if linha.startswith("CodFilial"):
            linha_cabecalho = i
            break

    if linha_cabecalho is None:
        raise ValueError(f"Cabeçalho CodFilial não encontrado em {caminho.name}")

    conteudo_limpo = "".join(linhas[linha_cabecalho:])
    df = pd.read_csv(StringIO(conteudo_limpo), sep=",", dtype=str)

    df.columns = df.columns.str.strip()
    df["CodFilial"] = df["CodFilial"].ffill()

    return df


def obter_meses_vendas(caminho: Path) -> list[str]:
    with open(caminho, encoding="cp1252") as arquivo:
        linhas = arquivo.readlines()

    for indice, linha in enumerate(linhas):
        if "MesDataVenda" in linha:
            linha_meses = linhas[indice + 1].strip()
            partes = linha_meses.split(",")

            meses = []

            for parte in partes:
                parte = parte.strip()

                if parte != "":
                    meses.append(parte.upper())

            return meses

    raise ValueError("Marcador MesDataVenda não encontrado no arquivo de vendas.")


def meses_por_modo(meses: list[str], modo: str) -> list[str]:
    if modo == "padrao":
        return meses[:4]

    if modo == "anual":
        return meses

    raise ValueError("Modo de relatório inválido. Use 'padrao' ou 'anual'.")


def nomes_colunas_vendas(meses: list[str], modo: str = "padrao") -> dict:
    meses_usados = meses_por_modo(meses, modo)

    colunas_qt = ["QtdeItem"] + [
        f"QtdeItem.{i}" for i in range(1, len(meses_usados))
    ]

    colunas_vlr = ["VlrLiqVenda"] + [
        f"VlrLiqVenda.{i}" for i in range(1, len(meses_usados))
    ]

    renomear = {}

    for i, mes in enumerate(meses_usados):
        mes = mes.upper()
        renomear[colunas_qt[i]] = f"QT {mes}"
        renomear[colunas_vlr[i]] = f"VLR {mes}"

    return renomear


def limpar_linhas_sem_produto(df: pd.DataFrame, coluna_produto: str) -> pd.DataFrame:
    df = df[df[coluna_produto].notna()]
    df = df[df[coluna_produto].str.strip() != ""]
    return df


def colunas_vendas_originais(meses: list[str], modo: str) -> list[str]:
    meses_usados = meses_por_modo(meses, modo)

    colunas = []

    for i in range(len(meses_usados)):
        if i == 0:
            colunas.append("QtdeItem")
            colunas.append("VlrLiqVenda")
        else:
            colunas.append(f"QtdeItem.{i}")
            colunas.append(f"VlrLiqVenda.{i}")

    return colunas


def gerar_lojas(
    estoque: pd.DataFrame,
    vendas: pd.DataFrame,
    meses: list[str],
    modo: str = "padrao",
) -> pd.DataFrame:
    lojas = estoque.merge(
        vendas,
        how="outer",
        left_on=["CodFilial", "CodProduto"],
        right_on=["CodFilial", "CódigoProduto"],
        suffixes=("_est", "_ven"),
    )

    estoque_cd = estoque[estoque["CodFilial"] == "900"][
        ["CodProduto", "QtEstoqueComercial"]
    ].copy()

    estoque_cd = estoque_cd.rename(
        columns={
            "CodProduto": "CD PROD",
            "QtEstoqueComercial": "QUANTIDADE ESTOQUE CD ATUAL",
        }
    )

    lojas["CD PROD"] = lojas["CodProduto"].fillna(lojas["CódigoProduto"])

    lojas = lojas.merge(
        estoque_cd,
        how="left",
        on="CD PROD",
    )

    lojas["PRODUTO"] = lojas["Descricao"].fillna(lojas["DescriçãoProduto"])
    lojas["Ean"] = lojas["EAN_est"].fillna(lojas["EAN_ven"]).fillna("S/ EAN Ativo")
    lojas["FABRICANTE"] = lojas["Fabricante_est"].fillna(lojas["Fabricante_ven"])
    lojas["Linha"] = lojas["Linha_est"].fillna(lojas["Linha_ven"])

    colunas_vendas = colunas_vendas_originais(meses, modo)

    lojas_final = lojas[
        [
            "CodFilial",
            "CD PROD",
            "PRODUTO",
            "Ean",
            "FABRICANTE",
            "Linha",
            "StatusProduto",
            "QtEstoqueComercial",
            "QUANTIDADE ESTOQUE CD ATUAL",
            "MediaF",
        ]
        + colunas_vendas
    ].copy()

    renomear_vendas = nomes_colunas_vendas(meses, modo)

    lojas_final = lojas_final.rename(
        columns={
            "CodFilial": "FILIAL",
            "StatusProduto": "Status Produtos",
            "QtEstoqueComercial": "QUANTIDADE ESTOQUE FILIAL ATUAL",
            "MediaF": "MEDIAF UN",
            **renomear_vendas,
        }
    )

    lojas_final = lojas_final.fillna(0)

    return lojas_final


def gerar_rede(
    estoque: pd.DataFrame,
    vendas: pd.DataFrame,
    meses: list[str],
    modo: str = "padrao",
) -> pd.DataFrame:
    estoque_lojas = estoque[estoque["CodFilial"] != "900"]
    estoque_cd = estoque[estoque["CodFilial"] == "900"]

    estoque_agrupado = estoque_lojas.groupby("CodProduto", as_index=False).agg(
        {"QtEstoqueComercial": "sum"}
    )

    estoque_agrupado = estoque_agrupado.rename(
        columns={"QtEstoqueComercial": "EST UN FILIAL GERAL"}
    )

    estoque_cd_base = estoque_cd[
        [
            "CodProduto",
            "EAN",
            "Descricao",
            "Fabricante",
            "Linha",
            "StatusProduto",
            "MediaF",
            "QtEstoqueComercial",
        ]
    ].copy()

    estoque_cd_base = estoque_cd_base.rename(
        columns={
            "CodProduto": "CD PROD",
            "EAN": "Ean",
            "Descricao": "PRODUTO",
            "Fabricante": "FABRICANTE",
            "StatusProduto": "Status Produtos",
            "MediaF": "MEDIAF UN GERAL",
            "QtEstoqueComercial": "EST UN CD GERAL",
        }
    )

    agregacoes = {}

    for coluna in colunas_vendas_originais(meses, modo):
        agregacoes[coluna] = "sum"

    vendas_agrupadas = vendas.groupby("CódigoProduto", as_index=False).agg(agregacoes)

    renomear_vendas = nomes_colunas_vendas(meses, modo)

    vendas_agrupadas = vendas_agrupadas.rename(
        columns={
            "CódigoProduto": "CD PROD",
            **renomear_vendas,
        }
    )

    rede = estoque_cd_base.merge(
        estoque_agrupado,
        how="outer",
        left_on="CD PROD",
        right_on="CodProduto",
    )

    rede = rede.merge(vendas_agrupadas, how="outer", on="CD PROD")

    rede["Ean"] = rede["Ean"].fillna("S/ EAN Ativo")

    colunas_finais = [
        "Ean",
        "CD PROD",
        "PRODUTO",
        "FABRICANTE",
        "Linha",
        "Status Produtos",
        "EST UN FILIAL GERAL",
        "EST UN CD GERAL",
        "MEDIAF UN GERAL",
    ]

    for mes in meses_por_modo(meses, modo):
        mes = mes.upper()
        colunas_finais.append(f"QT {mes}")
        colunas_finais.append(f"VLR {mes}")

    rede = rede[colunas_finais].copy()
    rede = rede.fillna(0)

    return rede


def formatar_excel(caminho_arquivo):
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    wb = load_workbook(caminho_arquivo)

    azul = PatternFill(
        fill_type="solid",
        start_color="1F4E78",
        end_color="1F4E78",
    )

    for ws in wb.worksheets:
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 30

        for cell in ws[1]:
            cell.font = Font(
                bold=True,
                color="FFFFFF",
            )

            cell.fill = azul

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        for coluna in ws.columns:
            largura = 0
            letra = get_column_letter(coluna[0].column)

            for celula in coluna:
                if celula.value is not None:
                    largura = max(
                        largura,
                        len(str(celula.value)),
                    )

            ws.column_dimensions[letra].width = min(
                largura + 2,
                50,
            )

        for coluna in ws.iter_cols(min_row=2):
            cabecalho = ws.cell(
                row=1,
                column=coluna[0].column,
            ).value

            if cabecalho is None:
                continue

            cabecalho = str(cabecalho).upper()

            for celula in coluna:
                if "VLR" in cabecalho or "VENDA LIQUIDA" in cabecalho:
                    celula.number_format = 'R$ #,##0.00'

                elif "QTDE" in cabecalho or "QT " in cabecalho or "ESTOQUE" in cabecalho:
                    celula.number_format = '#,##0'

                elif "MEDIAF" in cabecalho:
                    celula.number_format = '#,##0.00'

    wb.save(caminho_arquivo)


# ==========================
# CONFIGURAÇÃO DO RELATÓRIO
# ==========================

modo_relatorio = "anual"
# Opções:
# "padrao" = Padrão - 3 meses, ou seja, mês atual + 3 anteriores
# "anual" = todos os meses existentes no arquivo de vendas


# ==========================
# ARQUIVOS DE ENTRADA
# ==========================

pasta_entrada = Path("entrada")

arquivos_estoque = list(pasta_entrada.glob("*Estoque*.txt"))
arquivos_vendas = list(pasta_entrada.glob("*Vendas*.txt"))

if len(arquivos_estoque) == 0:
    raise FileNotFoundError("Nenhum arquivo de Estoque encontrado na pasta entrada.")

if len(arquivos_vendas) == 0:
    raise FileNotFoundError("Nenhum arquivo de Vendas encontrado na pasta entrada.")

if len(arquivos_estoque) > 1:
    raise ValueError(
        "Mais de um arquivo de Estoque encontrado. "
        "Deixe apenas um arquivo de Estoque na pasta entrada."
    )

if len(arquivos_vendas) > 1:
    raise ValueError(
        "Mais de um arquivo de Vendas encontrado. "
        "Deixe apenas um arquivo de Vendas na pasta entrada."
    )

arquivo_estoque = arquivos_estoque[0]
arquivo_vendas = arquivos_vendas[0]

meses_vendas = obter_meses_vendas(arquivo_vendas)

print(f"Arquivo de estoque encontrado: {arquivo_estoque.name}")
print(f"Arquivo de vendas encontrado: {arquivo_vendas.name}")
print(f"Meses encontrados: {meses_vendas}")
print(f"Modo do relatório: {modo_relatorio}")


# ==========================
# LEITURA DOS TXT
# ==========================

estoque = ler_txt_a_partir_do_cabecalho(arquivo_estoque)
estoque = limpar_linhas_sem_produto(estoque, "CodProduto")

vendas = ler_txt_a_partir_do_cabecalho(arquivo_vendas)
vendas = limpar_linhas_sem_produto(vendas, "CódigoProduto")


# ==========================
# CONVERSÃO NUMÉRICA
# ==========================

estoque["MediaF"] = estoque["MediaF"].apply(numero_br)
estoque["QtEstoqueComercial"] = estoque["QtEstoqueComercial"].apply(numero_br)

for coluna in colunas_vendas_originais(meses_vendas, modo_relatorio):
    vendas[coluna] = vendas[coluna].apply(numero_br)


# ==========================
# GERA RELATÓRIOS
# ==========================

lojas = gerar_lojas(estoque, vendas, meses_vendas, modo_relatorio)
rede = gerar_rede(estoque, vendas, meses_vendas, modo_relatorio)


# ==========================
# EXPORTA EXCEL
# ==========================

pasta_saida = Path("saida")
pasta_saida.mkdir(exist_ok=True)

fabricante = rede["FABRICANTE"].replace(0, pd.NA).dropna().iloc[0]
data_hoje = datetime.now().strftime("%d.%m")

nome_arquivo = f"Mapa de Vendas - {fabricante} {data_hoje}.xlsx"
arquivo_saida = pasta_saida / nome_arquivo

with pd.ExcelWriter(arquivo_saida, engine="openpyxl") as writer:
    rede.to_excel(writer, sheet_name="Rede", index=False)
    lojas.to_excel(writer, sheet_name="Lojas", index=False)

formatar_excel(arquivo_saida)


# ==========================
# CONFIRMAÇÃO
# ==========================

print("Rede gerada:", len(rede), "linhas")
print("Lojas geradas:", len(lojas), "linhas")
print()
print(f"Arquivo gerado com sucesso: {arquivo_saida}")